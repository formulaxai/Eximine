from django.shortcuts import render, HttpResponse, redirect,HttpResponseRedirect, get_object_or_404
from django.contrib.auth import authenticate, login as dj_login, logout
from django.contrib import messages 
from django.forms.utils import ErrorList
# from django.contrib.auth.models import User
from app.models import registrations, packages, faqs
# from verify_email.email_handler import send_verification_email
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_text, DjangoUnicodeDecodeError
from app.utils import generate_token
from django.views.generic import View, ListView
import threading
from django.urls import reverse
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from validate_email import validate_email
from django.contrib.auth.tokens import PasswordResetTokenGenerator
import stripe
from .serializers import RegisterSerializer
from rest_framework.views import APIView
from rest_framework import status
from django.http import JsonResponse
from django.contrib.auth.hashers import make_password,check_password
import xlwt
from app.forms import profile_form,importexportdata_Form
# from django.contrib.sites.models import Site
# from django.contrib.auth.hashers import make_password, check_password
from django.core.exceptions import ObjectDoesNotExist
# from .models import Membership, UserMembership, Subscription
# adding manualy subcribe library
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
from django.http import JsonResponse
from .models import *
# from .serializers import RegisterSerializer
from django.contrib import auth
import datetime
from datetime import timedelta
from datetime import datetime as dt
import requests
import json
from django.http import HttpResponseRedirect
from app.encript import hashPassword, verifyPassword
from django.utils.html import strip_tags
from django.core.mail import EmailMultiAlternatives
# import razorpay
from django.core.paginator import Paginator
from django.db.models import Q
from functools import reduce
import operator
import openpyxl
from openpyxl import Workbook
from html import unescape
import csv
import os
from django.contrib.admin.views.decorators import staff_member_required
import io
today = datetime.date.today()


# # Create your views here.
def index(request):
    data = packages.objects.all()
    faq_data = faqs.objects.all()
    context= {'data':data, 'faq_data':faq_data}
    return render(request, 'index.html',context)

def about(request):
    # return HttpResponse("this is aboutpage")
    return render(request, 'about.html')

def contact(request):
    # return HttpResponse('This is contactpage')
    return render(request, 'contact.html')

def dashboard(request):

    if 'id' not in request.session:
        messages.error(request,'Please login first!')
        return redirect("login")
    countries_object = countries.objects.all().distinct()
    print("countries_object", countries_object)
    data = registrations.objects.get(id=request.session['id'])
    return render(request, 'dashboard.html', {"data":data, 'countries_object': countries_object})

def register(request,pk=None):

    if 'id' in request.session:
        messages.success(request,"You are already Logged In!")
        return redirect("dashboard")
    
    packageid = pk

    if packages.objects.filter(packageid=pk).exists():
        data = packages.objects.get(packageid=pk)
    else:
        messages.error(request,'Please choose atleast one package to continue!')
        return HttpResponseRedirect('/')

    # if pk == '':
    #     packageid = request.POST['packageid']
    # elif pk == None:
    #     packageid = request.POST['packageid']
    # else:
    #     packageid = pk
    
    # if packageid == '':
    #     messages.error(request,'Please choose atleast one package to continue!')
    #     return HttpResponseRedirect('/')
    
    if request.method == 'POST':
        fullname = request.POST['fullname']
        password1 = hashPassword(request.POST['password1'])
        password2 = hashPassword(request.POST['password2'])
        email = request.POST['email']
        country_code = request.POST['country_code']
        phone_number = request.POST['phone_number']
        companyname = request.POST['companyname']
        user_type = request.POST['user_type']
        user_type_other = request.POST['user_type_other']
        
        package_details= packages.objects.get(packageid=packageid)

        current_Date = datetime.date.today()

        packageid=package_details.packageid
        plan_name=package_details.plan_name
        package_validity=package_details.package_validity
        price=package_details.price
        download_points=package_details.download_points
        download_points_total=package_details.download_points
        download_points_perday=package_details.download_points_perday
        download_points_perday_total=package_details.download_points_perday
        unlimited_data_access=package_details.unlimited_data_access
        unlimited_data_access_date=current_Date + timedelta(int(package_details.unlimited_data_access))
        searches=package_details.searches
        searches_total=package_details.searches
        work_spaces=package_details.work_spaces
        work_spaces_validity=package_details.work_spaces_validity
        work_spaces_validity_date=current_Date + timedelta(int(package_details.work_spaces_validity))
        work_spaces_deletion=package_details.work_spaces_deletion
        shipment_limit_workspaces=package_details.shipment_limit_workspaces
        addon_points_facility=package_details.addon_points_facility
        hot_products=package_details.hot_products
        hot_products_total=package_details.hot_products
        hot_products_interval=package_details.hot_products_interval
        hot_companies=package_details.hot_companies
        hot_companies_total=package_details.hot_companies
        hot_companies_interval=package_details.hot_companies_interval
        allowed_users=package_details.allowed_users
        allowed_users_total=package_details.allowed_users
        
        start_on = current_Date
        expire_on = current_Date + timedelta(int(package_details.package_validity))

        is_active = 0
        createtime = datetime.datetime.now().time()
        createdate = datetime.datetime.now().date()
        
        if password1==password2:
            if registrations.objects.filter(phone_number=phone_number).exists():
                messages.error(request,'Mobile Number is already in use.')
                return render(request,'register.html', {"data":data})
            elif registrations.objects.filter(email=email).exists():
                messages.error(request,'Email-Id is already in use.')
                return render(request,'register.html', {"data":data})
            else:
                user = registrations.objects.create(fullname=fullname,email=email,password=password1,country_code=country_code,phone_number=phone_number,companyname=companyname,user_type=user_type,user_type_other=user_type_other,packageid=packageid,plan_name=plan_name,package_validity=package_validity,price=price,download_points=download_points,download_points_total=download_points_total,download_points_perday=download_points_perday,download_points_perday_total=download_points_perday_total,unlimited_data_access=unlimited_data_access,unlimited_data_access_date=unlimited_data_access_date,searches=searches,searches_total=searches_total,work_spaces=work_spaces,work_spaces_validity=work_spaces_validity,work_spaces_validity_date=work_spaces_validity_date,work_spaces_deletion=work_spaces_deletion,shipment_limit_workspaces=shipment_limit_workspaces,addon_points_facility=addon_points_facility,hot_products=hot_products,hot_products_total=hot_products_total,hot_products_interval=hot_products_interval,hot_companies=hot_companies,hot_companies_total=hot_companies_total,hot_companies_interval=hot_companies_interval,allowed_users=allowed_users,allowed_users_total=allowed_users_total,start_on=start_on,expire_on=expire_on,is_active=is_active,createtime=createtime,createdate=createdate)

                # user.password =hashPassword(user.password)

                user.is_active = False
                user.save()

                # adding email authintication

                current_site = get_current_site(request)
                email_subject = 'Verify your email - Eximine'
                message = render_to_string('register-complete.html',
                {
                    'user':user,
                    'domain':current_site.domain,
                    'uid':urlsafe_base64_encode(force_bytes( user.pk)),
                    'token':generate_token.make_token(user),
                })
                
                email_message = EmailMessage(
                    email_subject,
                    message,
                    settings.EMAIL_HOST_USER,
                    [email],
                )
                email_message.content_subtype = "html"
                email_message.send()
                # EmailThread(email_message).start()
                messages.success(request, 'You have successfully registered with us. Please verify your email first and then login here.')
                return redirect('login')
        else:
            messages.error(request,'Password and confirm password should be same')    
            return render(request, 'register.html', {"data":data})
    else:
        return render(request,'register.html', {"data":data})

class registercomplete(View):
    def get(self, request, uidb64, token):
        try:
            uid=force_text(urlsafe_base64_decode(uidb64))
            user=registrations.objects.get(pk=uid)

        # except Exception as identifier:
        except(TypeError, ValueError, OverflowError, ObjectDoesNotExist):
            user = None

        if user is not None and generate_token.check_token(user,token):
            user.is_active=True
            user.save()
            messages.add_message(request, messages.INFO, 'Your account is activated now. You can login!')
            return redirect('login')
        
        messages.error(request,"There is some problem. Please try again!")
        return render(request,'login.html',status=401)

def login(request):
    
    if 'id' in request.session:
        messages.success(request,"You are already Logged In!")
        return redirect("dashboard")
    
    if request.method == 'POST':
        email = request.POST['email']
        password = hashPassword(request.POST['password'])
        remember = request.POST.get('remember')
        user=registrations.objects.filter(email=email,password=password)
        
        if user.exists():
            data = registrations.objects.get(email=email)
            if data.is_active == 1:
                if remember == "on":
                    request.session.set_expiry(settings.KEEP_LOGGED_DURATION)
                request.session['id'] = data.id
                messages.success(request, 'You are Logged In!')
                return redirect("dashboard")
            else:
                messages.error(request,'Your account is still Inactive. Please verify your email first.')
                return render(request, 'login.html')
        else:
            messages.error(request,'Please use valid credentials')
            return render(request, 'login.html')

    return render(request, 'login.html')

def logout(request):
    
    request.session.modified = True
    del request.session['id']
    request.session.flush()

    messages.success(request,"You have successfully Logged Out!")
    return HttpResponseRedirect('/')

def forgotpassword(request):
    
    if 'id' in request.session:
        messages.success(request,"You are already Logged In!")
        return redirect("dashboard")
    
    if request.method == 'POST':
        email = request.POST['email']

        if not validate_email(email):
            messages.error(request, 'Please enter a valid email')
            return render(request, 'forgot-password.html')
        
        registerCount = registrations.objects.filter(email=email).exists()
        
        if registerCount:
            user = registrations.objects.get(email=email)

            current_site = get_current_site(request)
            email_subject = 'Account Recovery - Eximine'
            message = render_to_string('forgot-password-email.html',
                {
                    'user':user,
                    'domain':current_site.domain,
                    'uid':urlsafe_base64_encode(force_bytes(user.pk)),
                    'token':generate_token.make_token(user)
                }
            )

            email_message = EmailMessage(
                email_subject,
                message,
                settings.EMAIL_HOST_USER,
                [email]
            )

            email_message.content_subtype = "html"
            email_message.send()

            messages.success(request,'We have sent an email with instructions on how to reset your password')
            return redirect("login")
        else:
            messages.error(request,'We are not able to recognize your email')
            return render(request, 'forgot-password.html')
    
    return render(request, 'forgot-password.html')

class forgotpasswordcomplete(View):
    def get(self, request, uidb64, token):
        context = {
            'uidb64': uidb64,
            'token': token
        }

        try:
            user_id = force_text(urlsafe_base64_decode(uidb64))

            user = registrations.objects.get(pk=user_id)

            if not generate_token.check_token(user, token):
                messages.info(
                    request, 'Password reset link, is invalid, please request a new one')
                return render(request, 'request-reset-email.html')

        except DjangoUnicodeDecodeError as identifier:
            messages.success(
                request, 'Invalid link')
            return render(request, 'request-reset-email.html')

        return render(request, 'forgot-password-complete.html', context)

    def post(self, request, uidb64, token):
        context = {
            'uidb64': uidb64,
            'token': token,
            'has_error': False
        }

        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        if len(password) < 6:
            messages.add_message(request, messages.ERROR,'passwords should be at least 6 characters')
            context['has_error'] = True
        if password != password2:
            messages.add_message(request, messages.ERROR,'Password and Confirm Password should be same!')
            context['has_error'] = True

        if context['has_error'] == True:
            return render(request, 'forgot-password-complete.html', context)

        try:
            user_id = force_text(urlsafe_base64_decode(uidb64))

            user = registrations.objects.get(pk=user_id)
            registrations.objects.filter(pk=user_id).update(password=hashPassword(password))

            messages.success(
                request, 'You have successfully reset your account password. Now you can login with your new password!')

            return redirect('login')

        except DjangoUnicodeDecodeError as identifier:
            messages.error(request, 'Something went wrong')
            return render(request, 'forgot-password-complete.html', context)

        return render(request, 'forgot-password-complete.html', context)

def myprofile(request):

    if 'id' not in request.session:
        messages.error(request,'Please login first!')
        return redirect("login")
    
    if request.method == 'POST':
        id = request.POST['id']
        fullname = request.POST['fullname']
        companyname = request.POST['companyname']
        user_type = request.POST['user_type']
        user_type_other = request.POST['user_type_other']
        
        modifytime = datetime.datetime.now().time()
        modifydate = datetime.datetime.now().date()

        registrations.objects.filter(id=id).update(fullname=fullname,companyname=companyname,user_type=user_type,user_type_other=user_type_other,modifytime=modifytime,modifydate=modifydate)
        messages.success(request, 'You have successfully updated your profile!')

    data = registrations.objects.get(id=request.session['id'])
    
    return render(request, 'myprofile.html', {"data":data})

def search(request):
    
    if 'id' not in request.session:
        messages.error(request,'Please login first!')
        return redirect("login")

    filter_names = ('HS_CODE',)

    filter_clauses = [Q(**{filter: request.GET[filter]})
                    for filter in filter_names
                    if request.GET.get(filter)]

    export_count = importexportdata.objects.filter(TYPE='EXPORT')
    if filter_clauses:
        export_count = export_count.filter(reduce(operator.and_, filter_clauses))
    export_count = export_count.count()
    if export_count == '':
        export_count = 0
    
    import_count = importexportdata.objects.filter(TYPE='IMPORT')
    if filter_clauses:
        import_count = import_count.filter(reduce(operator.and_, filter_clauses))
    import_count = import_count.count()
    if import_count == '':
        import_count = 0

    data_list = importexportdata.objects.all().order_by('-id')
    if filter_clauses:
        data_list = data_list.filter(reduce(operator.and_, filter_clauses))
    paginator = Paginator(data_list, 50) # Show 50 records per page.

    page_number = request.GET.get('page')
    all_data = paginator.get_page(page_number)

    data = registrations.objects.get(id=request.session['id'])
    
    return render(request, 'search.html', {'data':data, 'all_data':all_data, 'paginator':paginator, 'export_count':export_count, 'import_count':import_count})

def advancesearch(request):
    filter = {}
    if 'id' not in request.session:
        messages.error(request,'Please login first!')
        return redirect("login")
    select_country = request.GET.get('select_country')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # filter['COUNTRY'] = 'India'
    filter['COUNTRY'] = select_country
    if start_date and end_date:
        filter['DATE__range'] = [start_date, end_date]
    # DATE__range=["2020-11-15", "2020-11-20"]

    print("filter", filter)
    export_count = importexportdata.objects.filter(TYPE='EXPORT', **filter).count()
    if export_count == '':
        export_count = 0
    
    import_count = importexportdata.objects.filter(TYPE='IMPORT', **filter).count()
    if import_count == '':
        import_count = 0

    data_list = importexportdata.objects.filter(**filter).order_by('-id')
    paginator = Paginator(data_list, 50) # Show 50 records per page.

    page_number = request.GET.get('page')
    all_data = paginator.get_page(page_number)

    data = registrations.objects.get(id=request.session['id'])
    
    return render(request, 'search.html', {'data':data, 'all_data':all_data, 'paginator':paginator, 'export_count':export_count, 'import_count':import_count})
        
def users(request):
    return render(request, 'users.html')

def accessLog(request):
    return render(request, 'accessLog.html')

def favoriteshipment (request):
    return render(request, 'favoriteshipment.html')

def icons_icomoon(request):
    return render(request, 'icons_icomoon.html')

def mydownloads(request):
    return render(request, 'mydownloads.html')

def mynotifications(request):
    return render(request, 'mynotifications.html')

def myworkspace(request):
    return render(request, 'myworkspace.html')

def subscriptions(request):
    return render(request, 'subscriptions.html')

def tickets(request):
    return render(request, 'tickets.html')

def exportdata(request):
    """
    Downloads all data as Excel file with a single worksheet
    """
    data_queryset = importexportdata.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Import-Export Data'

    # Define the titles for columns
    columns = [
        'TYPE', 'DATE', 'MONTH', 'YEAR','HS_CODE', 'TWO_DIGIT','FOUR_DIGIT'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all datas
    for data in data_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            unescape(data.TYPE) if data.TYPE else '',
            unescape(data.DATE) if data.DATE else '',
            unescape(data.MONTH) if data.MONTH else '',
            unescape(data.YEAR) if data.YEAR else '',
            unescape(data.HS_CODE) if data.HS_CODE else '',
            unescape(data.TWO_DIGIT) if data.TWO_DIGIT else '',
            unescape(data.FOUR_DIGIT) if data.FOUR_DIGIT else '',
        ]
        print('row', row)
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


@staff_member_required
def upload_csv(request):
    
    # if request.method == 'GET':
    #     form = importexportdata_Form()
    #     return render(request, 'upload_csv.html', {'form':form})
    
    if request.method == "POST":
        country = request.POST.get('country')
        form = importexportdata_Form(request.POST, request.FILES)
        if form.is_valid():
            csv_file = importexportdata()
            csv_file.csv_file = form.cleaned_data['csv_file']
            # if csv_file:
            #     ext = csv_file.name.split('.')[-1]
            #     if ext != 'csv':
            #         raise forms.ValidationError('File Type is not Supported')

            #     return csv_file
            csv_file = io.TextIOWrapper(form.cleaned_data['csv_file'])
            reader = csv.reader(csv_file, delimiter=',')
                                        # reader = csv.DictReader(StringIO(csv_file.read().decode('utf-8')), delimiter=';')
            row =0
            for data in reader:
                if row==0:
                    headers = data
                    row = row + 1
                else:
                    new_data ={}
                    for i in range(len(headers)):
                        new_data[headers[i]] = data[i]

                    # new_data1 = importexportdata()
                    new_data1=importexportdata()
                    
                    new_data1.__dict__.update(new_data)
                    new_data1.save()
                    # print(new_data1.id)
                    importexportdata.objects.filter(id=new_data1.id).update(COUNTRY=country)
                    row= row + 1

        country_data = countries.objects.all()
        messages.success(request,'Data Uploaded Successfully!')
        return render(request,'upload_csv.html', {"country_data":country_data})

    # else:
    #     form = importexportdataForm(request.POST)
    
    country_data = countries.objects.all()
    # print(country_data)
    return render(request, 'upload_csv.html', {'country_data':country_data})